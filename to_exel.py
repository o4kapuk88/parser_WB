import os
import openpyxl


class ExcelManager:
    def __init__(self, filename, headers):
        self.filename = filename
        self.headers = headers

    def get_or_create_sheet_with_headers(self, wb):
        sheet_name = 'List_1'
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(self.headers)
            return ws

    def save_to_excel(self, data):
        if not os.path.isfile(self.filename):
            wb = openpyxl.Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)
            ws = self.get_or_create_sheet_with_headers(wb)
        else:
            wb = openpyxl.load_workbook(self.filename)
            ws = self.get_or_create_sheet_with_headers(wb)

        start_row = ws.max_row + 1

        for row_data in data:
            for col, value in enumerate(row_data, start=1):
                ws.cell(row=start_row, column=col, value=value)
            start_row += 1

        wb.save(filename=self.filename)





